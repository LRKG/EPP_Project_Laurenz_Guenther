import pandas as pd
from bld.project_paths import project_paths_join as ppj
import nltk
from openpyxl import load_workbook
import random
from datetime import datetime
import matplotlib.pyplot as plt


# Load the training data.
wb=load_workbook(ppj("OUT_DATA", "training.xlsx"))
sheet=wb.active
max_row=sheet.max_row

# Create a tuple containing for each article the words as an array and for islamist
# terrorism an indicator.
tuple = []
for row_id in range(2, max_row+1): 
    tuple.append(([x.strip() for x in str(sheet.cell(row=row_id, column=1).value).split(';')], str(sheet.cell(row=row_id, column=4).value)))

# Randomise the order of the articles.
random.shuffle(tuple)

# Generate an array containing all words of all articles I classified.
all_words = []
for article in range(0, len(tuple)):
    for w in tuple[article][0]:
        all_words.append(w.lower())

# Generate a frequency distribution of those words i.e. count for each how often it occured.      
all_words = nltk.FreqDist(all_words)    
# Make a list with number of occurences for the words that occur most often.
word_features = list(all_words.keys())[:10000]
    


def find_features(article):
    '''
    Tell for each of the 3000 words that occur most often
    whether it occurs in a given article or not.
    Input: Id of an article.
    Output: Dictionary containing as keys the 3000 words that occur most often
    in all articles. Values are boolean and indicate whether the given article
    contains the word or not (True/False)
    '''
    words = set(article)    
    features = {}
    for w in word_features:
        features[w] = (w in words)        
    return features


        
# Generate a tuple consisting of a dictionary indicating which of the 3000 words
# that generally occur most often is contained in each article and one indicator
# for the kind of terrorism, for each kind of terrorism. 
feature_sets = []
for i in range(0, len(tuple)-1):
    feature_sets.append((find_features(tuple[i][0]), str(sheet.cell(row=i+2, column=4).value)))

training_set = feature_sets[:int(len(feature_sets)/2)]
testing_set = feature_sets[int(len(feature_sets)/2):]

classifier = nltk.NaiveBayesClassifier.train(training_set)
#print("Original naive Bayes Algorithm Accuracy in percent:", (nltk.classify.accuracy(classifier, testing_set)*100))
#classifier.show_most_informative_features(10)
#classifier.labels()



# Load the data for the prediction.
wb_2=load_workbook(ppj("OUT_DATA", "all_articles.xlsx"))
sheet_2=wb_2.active
max_row_2=sheet_2.max_row

# Create a tuple containing for each article the words as an array.
tuple_2 = []
for row_id in range(2, max_row_2+1): 
    tuple_2.append(([x.strip() for x in str(sheet_2.cell(row=row_id, column=2).value).split(';')]))

# Generate a tuple consisting of a dictionary indicating which of the 3000 words
# that generally occur most often is contained in each article.
feature_sets_2 = []
for i in range(0, int(len(tuple_2)-1)):
    feature_sets_2.append((find_features(tuple_2[i])))

# Classify the articles with ex-ante unknown label.
classifications_islam = classifier.classify_many(feature_sets_2)
#print(classifications_islam)


# Save the number of articles classified as being about islamist terrorism and
# the corresponding years in a dictionary.
other_year = 0
year_old = int(datetime.strptime(sheet_2.cell(row=0 + 2, column=1).value, '%Y/%m/%d').year)
year_new = 0
islam_year = 0
dict = {'year': [], "islam": []}
for row_id in range(0, max_row_2 -2):     
    if classifications_islam[row_id] == "islam":      
        islam_year =islam_year + 1     
    if (datetime.strptime(sheet_2.cell(row=row_id + 2, column=1).value, '%Y/%m/%d').year > year_old):
        year_new = int(datetime.strptime(sheet_2.cell(row=row_id + 2, column=1).value, '%Y/%m/%d').year)
        
        dict["islam"].append(islam_year)         
        dict["year"].append(year_old)
        islam_year = 0
        year_old = year_new
 
dict["islam"].append(islam_year)         
dict["year"].append(year_old)       
       
    
# Comvert the data to a pandas data frame.
df = pd.DataFrame(data=dict)

# Plot the data as a line-chart.
plt.plot( 'year', 'islam', data=df, marker='', color='red', linewidth=2, linestyle='dashed', label="number of reports on islamist terrorist attacks")
#plt.plot( 'year', 'other', data=df, marker='', color='red', linewidth=2, linestyle='_ _', label="toto")
plt.legend()

# Save the plot.
plt.savefig(ppj("OUT_FIGURES", "line_chart.pdf"), delimiter=",")




