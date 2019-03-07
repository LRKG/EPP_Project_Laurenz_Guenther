from bld.project_paths import project_paths_join as ppj
from urllib.request import urlopen as ureq
from bs4 import BeautifulSoup as soup
import xlsxwriter
from openpyxl import load_workbook
from nltk.tokenize import word_tokenize 
import random



# Load data that was collected manually and indicates for some dates the 
# headlines of articles concerned with terrorism.
filepath=ppj("IN_DATA", "Bild_Training.xlsx")
wb=load_workbook(filepath)
sheet=wb.active
max_row=sheet.max_row


# Create an empty dictionary that will be filled later.
#dict = {'words': [], "right": [], "left": [], "islam": [], "other": []}
dict_training = {'words': [], "islam": []}

# Create arrays for the names of articles about different kinds of terrorism. 
#right_article_names_array = []
#left_article_names_array = []
islam_article_names_array = []
#other_article_names_array = []

# Create an empty dictionary that will be filled later.
dict_prediction = {'words': [], 'date': []}


# Loop over all dates in the raw-excel file.
for i_row in range(0, max_row): 
    try:
        # Skip a date if it is not indicated as containing articles about terrorism 
        # otherwise extract the titles of the articles about terrorism and save
        # them in the respective array.
        if_indicator = 0
        '''
        if not sheet.cell(row=i_row,column=2).value is None:
            if float(sheet.cell(row=i_row,column=2).value) > 0:      
                if_indicator = if_indicator + 1
                right_article_names_array = [x.strip() for x in str(sheet.cell(row=i_row,column=3).value).split(';')]
            else:
                right_article_names_array = []          
        else:
            right_article_names_array = []
            
         
        if not sheet.cell(row=i_row,column=4).value is None:
            if float(sheet.cell(row=i_row,column=4).value) > 0:
                    if_indicator = if_indicator + 1
                    left_article_names_array = [x.strip() for x in str(sheet.cell(row=i_row,column=5).value).split(';')] 
            else:
                left_article_names_array = []          
        else:
            left_article_names_array = []
            
        if not sheet.cell(row=i_row,column=8).value is None:
            if float(sheet.cell(row=i_row,column=8).value) > 0:
                if_indicator = if_indicator + 1
                other_article_names_array = [x.strip() for x in str(sheet.cell(row=i_row,column=9).value).split(';')]
            else:
                other_article_names_array = []           
        else:
            other_article_names_array = []          
        ''' 
               
        if not sheet.cell(row=i_row,column=6).value is None:
            if float(sheet.cell(row=i_row,column=6).value) > 0:
                if_indicator = if_indicator + 1
                islam_article_names_array = [x.strip() for x in str(sheet.cell(row=i_row,column=7).value).split(';')]   
            else:
                islam_article_names_array = []            
        else:
            islam_article_names_array = []

        r1 = random.randint(1, 900) 
        if (if_indicator > 0) or (if_indicator == 0 and r1 < 2):
            try:
                # Get the date from the excel-file and bring it in the right form to scrape.
                date = list(str(sheet.cell(row=i_row,column=1).value))             
                    
                del date[10:]
                date[4] = "/"
                   
                if(float(date[5]) == 0):
                    del date[5]
                    date[6] = "/"
                    if(float(date[7]) == 0):
                        del date[7]
                else:
                    date[7] = "/"
                    if(float(date[8]) == 0):
                        del date[8]
            
                date[4] = "/"
                date = "".join(date)
                #print(date)
                # Use the date to generate the url for the archive for each day.
                url = "https://www.bild.de/archive/%s/index.html"% (date)
                #print(url)
                
                # Grabb the page and save its content in a variable,then close the site.
                website = ureq(url)
                website_html = website.read()
                website.close()
                
                # Disentangle useful and useless content from the page.
                website_soup = soup(website_html, "html.parser")
                
                # Find all containers containing links to articles of one page.
                containers = website_soup.findAll("p") 
                
                
                # Loop over these links to single articles.
                # Begn at 8 as the first 8 contaiers are always reserved for uninteresting stuff.     
                for i in range(8, len(containers)):   
                    # Tell me at which date and which article you are.
                    #print("date=", date, "......", "container=", i )
                                   
                    # Access all links to individual article for a day.
                    article = website_soup.findAll("p")[i].a["href"]
                    article_url = "https://www.bild.de"+article 
                    #print(article_url)
                    
                    # Download the webpage for an individual article.
                    article_website = ureq(article_url) 
                    article_website_html = article_website.read()
                    article_website.close()
                    article_website_soup = soup(article_website_html, "html.parser")
                
                    try:
                        # Try to access the article. If this is not possible
                        # or the article is empty, go on.
                    
                        # Access the headline of the article. The encoding can
                        # differ by article.  
                        try: 
                            headline = article_website_soup.find("div",{"class":"content s12 articlemdot "}).find("span",{"class":"headline"}).text
                        except:
                            try:
                                headline = article_website_soup.find("div",{"class":"content s10 article "}).find("span",{"class":"headline"}).text
                            except:
                                headline = article_website_soup.find("div",{"class":"content s12 article "}).find("span",{"class":"headline"}).text
                           
                        #right_indicator = 0
                        #left_indicator = 0
                        islam_indicator = 0
                        #other_indicator = 0
                        terror_indicator = 0
                        
                        for i in range(0, len(islam_article_names_array)):                       
                            if (islam_article_names_array[i] == headline):
                                islam_indicator = islam_indicator + 1
                                terror_indicator = terror_indicator + 1
                            else:
                                pass
                        
                        '''
                        for i in range(0, len(right_article_names_array)):
                            if (right_article_names_array[i] == headline):
                                right_indicator = right_indicator + 1
                                terror_indicator = terror_indicator + 1
                            else:
                                pass
                           
                        for i in range(0, len(left_article_names_array)):                      
                            if (left_article_names_array[i] == headline):                         
                                left_indicator = left_indicator + 1
                                terror_indicator = terror_indicator + 1
                            else:
                              pass
                          
                        for i in range(0, len(other_article_names_array)):
                            if (other_article_names_array[i] == headline):
                                other_indicator = other_indicator + 1
                                terror_indicator = terror_indicator + 1
                            else:
                                pass
                        '''
                        
                        # If the article is non about terrorism at all, only save
                        # the article with a low probability to keep the sample of 
                        # articles balanced.
                        scrape_indicator = 0
                        r = random.randint(1, 200)

                        if (terror_indicator == 0):
                            if r < 2:
                                scrape_indicator = scrape_indicator + 1
                            else:
                                pass                  
                        else:
                            scrape_indicator = scrape_indicator + 1
                            
                        
                        if scrape_indicator == 1:
                                               
                            # Access all text-parts of the article. The encoding can differ by article.
                            try:
                                article_containers = article_website_soup.find("div",{"class":"txt clearfix"}).findAll('p')
                            except:
                                article_containers = article_website_soup.find("div",{"class":"txt"}).findAll('p')
                                   
                            # Access the kicker (above the heading) of the article. The encoding can differ by article.
                            try: 
                                kicker = article_website_soup.find("div",{"class":"content s12 articlemdot "}).find("span",{"class":"kicker"}).text
                            except:
                                try:
                                    kicker = article_website_soup.find("div",{"class":"content s10 article "}).find("span",{"class":"kicker"}).text
                                except:
                                    kicker = article_website_soup.find("div",{"class":"content s12 article "}).find("span",{"class":"kicker"}).text
                                                    
                            
                            # Combine all three parts as article text.
                            article_text = kicker +" "+ headline + " "      
                            for i in range(0, len(article_containers)):
                                article_text = article_text + article_containers[i].text
                              
                            # Calculate the length of the article.
                            #article_length = len(article_text.split())
                            
                            # Split the text by words and write the array of words
                            # to the dictionary.
                            tokenized_text = word_tokenize(article_text)
                            dict_training["words"].append(tokenized_text)
                            
                            
                            # Check whether the headline of the scraped article 
                            # matches one headline in the articles
                            # If so, write in the dictionary that the article is 
                            # about the respective form of terrorism.
                            if islam_indicator > 0:
                                dict_training["islam"].append("islam")
                            else:
                                dict_training["islam"].append("non-islam")
                                                  
                            '''
                            if right_indicator > 0:
                                dict["right"].append("right")
                            else:
                                dict["right"].append("non-right")
                                                     
                            if left_indicator > 0:                       
                                dict["left"].append("left")
                            else:
                                dict["left"].append("non-left")
                           
                            if other_indicator > 0:
                                dict["other"].append("other")
                            else:
                                dict["other"].append("non-other")
                            ''' 

                            # Split the text by words and write the array of words to the dictionary.
                            tokenized_text = word_tokenize(article_text)
                            dict_prediction["words"].append(tokenized_text)
                            dict_prediction["date"].append(date)
                

                    except:
                        pass
            except:
                pass      
        else:
            pass
    except:
        pass
                  
       
            

# Create a workbook for the output and add a worksheet.
workbook = xlsxwriter.Workbook(ppj("OUT_DATA", "all_articles.xlsx")) 
worksheet = workbook.add_worksheet()

# Write the eadings for the variables to it.
worksheet.write(0, 0, "date")   
worksheet.write(0, 1, "words")  
        

# To write the words-array in a compact form, make it a string.
words_strings = []
for i in range(0, len(dict_prediction["words"])): 
    words_strings.append(';'.join(dict_training["words"][i]))

# Loop over all articles in the dictionary and save the word-strings in the excel-sheet.
for row_id in range(0, len(dict_prediction["words"])): 
    worksheet.write(row_id + 1, 1, words_strings[row_id])    
    worksheet.write(row_id + 1, 0, str(dict_prediction["date"][row_id]) )        
        
workbook.close()



# Create a workbook for the output and add a worksheet.
workbook = xlsxwriter.Workbook(ppj("OUT_DATA", "training.xlsx")) 
worksheet = workbook.add_worksheet()

# Write the eadings for the variables to it.
worksheet.write(0, 0, "words")          
#worksheet.write(0, 1, "rightwing_indicator")
#worksheet.write(0, 2, "leftwing_indicator")
worksheet.write(0, 3, "islam_indicator")
#worksheet.write(0, 4, "other_indicator")  

# To write the words-array in a compact form, make it a string.
words_strings = []
for i in range(0, len(dict_training["words"])): 
    words_strings.append(';'.join(dict_training["words"][i]))

# Loop over all articles in the dictionary and save the word-strings and
# kind-of-terrorism indicators in the excel-sheet.
for row_id in range(1, len(dict_training["words"])):
    worksheet.write(row_id, 0, words_strings[row_id])          
    #worksheet.write(row_id, 1, dict["right"][row_id])
    #worksheet.write(row_id, 2, dict["left"][row_id])
    worksheet.write(row_id, 3, dict_training["islam"][row_id])
    #worksheet.write(row_id, 4, dict["other"][row_id])
        
workbook.close()

